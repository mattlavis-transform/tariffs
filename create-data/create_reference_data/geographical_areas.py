import psycopg2
import sys
import os
import xmlschema
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.objects as o
from common.geographical_area import geographical_area
from common.application import application

app = o.app
app.getTemplates()

try:
	profile = sys.argv[1]
except:
	profile = "geographical_areas"

fname = os.path.join(app.SOURCE_DIR, profile + ".xlsx")
wb = load_workbook(filename=fname, read_only=True)
ws = wb['Updated']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	GEOGRAPHICAL_AREA_SID   = ws.cell(row = i, column = 1).value
	GEOGRAPHICAL_AREA_ID    = ws.cell(row = i, column = 2).value
	DESCRIPTION             = ws.cell(row = i, column = 3).value

	obj = geographical_area(GEOGRAPHICAL_AREA_SID, GEOGRAPHICAL_AREA_ID, "", DESCRIPTION, "update")
	app.geographical_area_list.append(obj)

ws = wb['New']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	GEOGRAPHICAL_AREA_SID   = ws.cell(row = i, column = 1).value
	GEOGRAPHICAL_AREA_ID    = ws.cell(row = i, column = 2).value
	DESCRIPTION             = ws.cell(row = i, column = 3).value
	GEOGRAPHICAL_AREA_CODE  = ws.cell(row = i, column = 4).value

	""" AREA CODES
	0 Country
	1 Geographical area group
	2 Region
	"""

	obj = geographical_area(GEOGRAPHICAL_AREA_SID, GEOGRAPHICAL_AREA_ID, GEOGRAPHICAL_AREA_CODE, "", DESCRIPTION, "insert")
	app.geographical_area_list.append(obj)

env = app.envelope_XML
env = env.replace("{ENVELOPE_ID}", str(app.base_envelope_id))
out = ""
for obj in app.geographical_area_list:
	obj.writeXML(app)
	out += obj.xml

out = env.replace("{BODY}", out)
filename = os.path.join(app.XML_DIR, profile + ".xml")
f = open(filename, "w", encoding="utf-8")
f.write(out)
f.close()

schema_path = os.path.join(app.SCHEMA_DIR, "envelope.xsd")
my_schema = xmlschema.XMLSchema(schema_path)

try:
	if my_schema.is_valid(filename):
		print ("The file validated successfully.")
	else:
		print ("The file did not validate.")
except:
	print ("The file did not validate and crashed the validator.")
app.set_config()
