import psycopg2
import sys
import os
import xmlschema
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.objects as o
from common.membership import membership
from common.application import application

app = o.app
app.getTemplates()

fname = os.path.join(app.SOURCE_DIR, "memberships.xlsx")
wb = load_workbook(filename=fname, read_only=True)
ws = wb['New']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	PARENT_SID	= ws.cell(row = i, column = 1).value
	PARENT_ID	= ws.cell(row = i, column = 2).value
	CHILD_SID	= ws.cell(row = i, column = 3).value
	CHILD_ID	= ws.cell(row = i, column = 4).value
	START_DATE	= ws.cell(row = i, column = 5).value
	END_DATE	= ws.cell(row = i, column = 6).value
	UPDATE_TYPE	= ws.cell(row = i, column = 7).value

	obj = membership(PARENT_SID, PARENT_ID, CHILD_SID, CHILD_ID, START_DATE, END_DATE, UPDATE_TYPE)
	app.membership_list.append(obj)

env = app.envelope_XML
env = env.replace("{ENVELOPE_ID}", str(app.base_envelope_id))
out = ""
for obj in app.membership_list:
	obj.writeXML(app)
	out += obj.xml

out = env.replace("{BODY}", out)
filename = os.path.join(app.XML_DIR, "memberships.xml")
f = open(filename, "w", encoding="utf-8")
f.write(out)
f.close()

schema_path = os.path.join(app.SCHEMA_DIR, "envelope.xsd")
my_schema = xmlschema.XMLSchema(schema_path)

try:
	if my_schema.is_valid(filename):
		print ("The file validated successfully.")
	else:
		print ("The file did not validate.")
except:
	print ("The file did not validate and crashed the validator.")
app.set_config()
