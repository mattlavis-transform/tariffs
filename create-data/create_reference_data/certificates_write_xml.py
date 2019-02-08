import psycopg2
import sys
import os
import xmlschema
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.objects as o
from common.certificate import certificate
from common.application import application

app = o.app
app.getTemplates()

fname = os.path.join(app.SOURCE_DIR, "certificates.xlsx")
wb = load_workbook(filename=fname, read_only=True)

"""
ws = wb['Updated']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	CERTIFICATE_TYPE_CODE   = ws.cell(row = i, column = 1).value
	CERTIFICATE_CODE        = ws.cell(row = i, column = 2).value
	DESCRIPTION_OLD         = ws.cell(row = i, column = 3).value
	DESCRIPTION             = ws.cell(row = i, column = 4).value
	obj = certificate(CERTIFICATE_TYPE_CODE, CERTIFICATE_CODE, DESCRIPTION_OLD, DESCRIPTION, "update")
	app.certificates_list.append(obj)
"""
ws = wb['New']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	CERTIFICATE_TYPE_CODE   			= ws.cell(row = i, column = 1).value
	CERTIFICATE_CODE        			= ws.cell(row = i, column = 2).value
	DESCRIPTION             			= ws.cell(row = i, column = 3).value
	CERTIFICATE_DESCRIPTION_PERIOD_SID	= ws.cell(row = i, column = 4).value

	obj = certificate(CERTIFICATE_TYPE_CODE, CERTIFICATE_CODE, "", DESCRIPTION, CERTIFICATE_DESCRIPTION_PERIOD_SID, "insert")
	app.certificates_list.append(obj)

env = app.envelope_XML
env = env.replace("{ENVELOPE_ID}", str(app.base_envelope_id))
out = ""
for obj in app.certificates_list:
	obj.writeXML(app)
	out += obj.xml

out = env.replace("{BODY}", out)
filename = os.path.join(app.XML_DIR, "certificates.xml")
f = open(filename, "w", encoding="utf-8") 
f.write(out)
f.close()

schema_path = os.path.join(app.SCHEMA_DIR, "envelope.xsd")
my_schema = xmlschema.XMLSchema(schema_path)
try:
	if my_schema.is_valid(filename):
		print ("The file validated successfully.")
	else:
		print ("The file did not validate.")
except:
	print ("The file did not validate and crashed the validator.")
app.set_config()
