import psycopg2
import sys
import os
import xmlschema
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.objects as o
from common.footnote import footnote
from common.application import application

app = o.app
app.getTemplates()
app.get_profile("footnotes")

fname = os.path.join(app.SOURCE_DIR, "footnotes.xlsx")
wb = load_workbook(filename=fname, read_only=True)
ws = wb['Updated']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	FOOTNOTE_TYPE_ID = ws.cell(row = i, column = 1).value
	FOOTNOTE_ID      = ws.cell(row = i, column = 2).value
	DESCRIPTION_OLD  = ws.cell(row = i, column = 3).value
	DESCRIPTION      = ws.cell(row = i, column = 4).value

	f = footnote(FOOTNOTE_TYPE_ID, FOOTNOTE_ID, DESCRIPTION, DESCRIPTION_OLD, "update")
	app.footnotes_list.append(f)

ws = wb['New']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	FOOTNOTE_TYPE_ID = ws.cell(row = i, column = 1).value
	FOOTNOTE_ID      = ws.cell(row = i, column = 2).value
	DESCRIPTION      = ws.cell(row = i, column = 3).value

	f = footnote(FOOTNOTE_TYPE_ID, FOOTNOTE_ID, DESCRIPTION, "", "insert")
	app.footnotes_list.append(f)

env = app.envelope_XML
env = env.replace("{ENVELOPE_ID}", str(app.base_envelope_id))
out = ""
for f in app.footnotes_list:
	f.writeXML(app)
	out += f.xml

out = env.replace("{BODY}", out)
filename = os.path.join(app.XML_DIR, "footnotes.xml")
f = open(filename, "w", encoding="utf-8")
f.write(out)
f.close()

schema_path = os.path.join(app.SCHEMA_DIR, "envelope.xsd")
my_schema = xmlschema.XMLSchema(schema_path)
try:
	if my_schema.is_valid(filename):
		print ("The file validated successfully.")
	else:
		print ("The file did not validate.")
except:
	print ("The file did not validate and crashed the validator.")
app.set_config()
