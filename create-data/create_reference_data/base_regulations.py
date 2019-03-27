import psycopg2
import sys
import os

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.objects as o
from common.base_regulation import base_regulation
from common.application import application

app = o.app
app.d("Writing regulation XML", False)
app.getTemplates()
app.get_profile("regulations")

app.d("Reading Excel source file")
wb = load_workbook(filename=app.excel_source, read_only=True, data_only=True)

ws = wb['New']
row_count = ws.max_row
col_count = ws.max_column
filename = os.path.join(app.XML_DIR, app.xml_file)

for i in range(2, row_count + 1):
	BASE_REGULATION_ID  = ws.cell(row = i, column = 1).value
	VALIDITY_START_DATE = ws.cell(row = i, column = 2).value
	REGULATION_GROUP_ID = ws.cell(row = i, column = 3).value
	LEGISLATION_ID		= ws.cell(row = i, column = 4).value
	URL    				= ws.cell(row = i, column = 5).value
	INFORMATION_TEXT    = ws.cell(row = i, column = 6).value

	obj = base_regulation(BASE_REGULATION_ID, VALIDITY_START_DATE, REGULATION_GROUP_ID, LEGISLATION_ID, URL, INFORMATION_TEXT, "insert")
	app.base_regulations_list.append(obj)

ws = wb['Updated']
row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	BASE_REGULATION_ID  = ws.cell(row = i, column = 1).value
	VALIDITY_START_DATE = ws.cell(row = i, column = 2).value
	REGULATION_GROUP_ID = ws.cell(row = i, column = 3).value
	LEGISLATION_ID		= ws.cell(row = i, column = 4).value
	URL    				= ws.cell(row = i, column = 5).value
	INFORMATION_TEXT    = ws.cell(row = i, column = 6).value

	obj = base_regulation(BASE_REGULATION_ID, VALIDITY_START_DATE, REGULATION_GROUP_ID, LEGISLATION_ID, URL, INFORMATION_TEXT, "update")
	app.base_regulations_list.append(obj)

app.d("Writing XML file to " + filename)
env = app.envelope_XML
env = env.replace("{ENVELOPE_ID}", str(app.base_envelope_id))
out = ""
for obj in app.base_regulations_list:
	obj.writeXML(app)
	out += obj.xml

out = env.replace("{BODY}", out)
f = open(filename, "w", encoding="utf-8")
f.write(out)
f.close()

app.validate(filename)
app.set_config()
