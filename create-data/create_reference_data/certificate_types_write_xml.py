import psycopg2
import sys
import os
import xmlschema
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.objects as o
from common.certificate_type import certificate_type
from common.application import application

app = o.app
app.getTemplates()

fname = os.path.join(app.SOURCE_DIR, "certificate_types.xlsx")
wb = load_workbook(filename=fname, read_only=True)
ws = wb['Updated']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	CERTIFICATE_TYPE_ID = ws.cell(row = i, column = 1).value
	DESCRIPTION             = ws.cell(row = i, column = 2).value

	act = certificate_type(CERTIFICATE_TYPE_ID, DESCRIPTION, "update")
	app.certificate_type_list.append(act)

ws = wb['New']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	CERTIFICATE_TYPE_ID = ws.cell(row = i, column = 1).value
	DESCRIPTION             = ws.cell(row = i, column = 2).value

	f = certificate_type(CERTIFICATE_TYPE_ID, DESCRIPTION, "insert")
	app.certificate_type_list.append(f)

env = app.envelope_XML
env = env.replace("{ENVELOPE_ID}", str(app.base_envelope_id))
out = ""
for obj in app.certificate_type_list:
	obj.writeXML(app)
	out += obj.xml

out = env.replace("{BODY}", out)
filename = os.path.join(app.XML_DIR, "certificate_types.xml")
f = open(filename, "w", encoding="utf-8") 
f.write(out)
f.close()

schema_path = os.path.join(app.SCHEMA_DIR, "envelope.xsd")
my_schema = xmlschema.XMLSchema(schema_path)
try:
	if my_schema.is_valid(filename):
		print ("The file validated successfully.")
	else:
		print ("The file did not validate.")
except:
	print ("The file did not validate and crashed the validator.")
app.set_config()
