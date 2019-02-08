import psycopg2
import sys
import os
import xmlschema
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.objects as o
from common.additional_code_type import additional_code_type
from common.application import application

app = o.app
app.getTemplates()

fname = os.path.join(app.SOURCE_DIR, "additional_code_types.xlsx")
wb = load_workbook(filename=fname, read_only=True)
ws = wb['Updated']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	ADDITIONAL_CODE_TYPE_ID = ws.cell(row = i, column = 1).value
	DESCRIPTION             = ws.cell(row = i, column = 2).value
	APPLICATION_CODE        = "n/a"

	act = additional_code_type(ADDITIONAL_CODE_TYPE_ID, DESCRIPTION, APPLICATION_CODE, "update")
	app.additional_code_type_list.append(act)

ws = wb['New']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	ADDITIONAL_CODE_TYPE_ID = ws.cell(row = i, column = 1).value
	DESCRIPTION             = ws.cell(row = i, column = 2).value
	APPLICATION_CODE        = ws.cell(row = i, column = 3).value

	f = additional_code_type(ADDITIONAL_CODE_TYPE_ID, DESCRIPTION, APPLICATION_CODE, "insert")
	app.additional_code_type_list.append(f)

env = app.envelope_XML
env = env.replace("{ENVELOPE_ID}", str(app.base_envelope_id))
out = ""
for obj in app.additional_code_type_list:
	obj.writeXML(app)
	out += obj.xml

out = env.replace("{BODY}", out)
filename = os.path.join(app.XML_DIR, "additional_code_types.xml")
f = open(filename, "w", encoding="utf-8") 
f.write(out)
f.close()

schema_path = os.path.join(app.SCHEMA_DIR, "envelope.xsd")
my_schema = xmlschema.XMLSchema(schema_path)
try:
	if my_schema.is_valid(filename):
		print ("The file validated successfully.")
	else:
		print ("The file did not validate.")
except:
	print ("The file did not validate and crashed the validator.")
app.set_config()
