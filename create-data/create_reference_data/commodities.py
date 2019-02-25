import psycopg2
import sys
import os
import xmlschema
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.objects as o
from common.goods_nomenclature import goods_nomenclature
from common.application import application

app = o.app
app.getTemplates()

fname = os.path.join(app.SOURCE_DIR, "commodities.xlsx")
wb = load_workbook(filename=fname, read_only=True)
ws = wb['Updated']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	GOODS_NOMENCLATURE_ITEM_ID					= ws.cell(row = i, column = 1).value
	PRODUCTLINE_SUFFIX          				= ws.cell(row = i, column = 2).value
	DESCRIPTION         	    				= ws.cell(row = i, column = 3).value
	GOODS_NOMENCLATURE_SID						= ws.cell(row = i, column = 4).value
	GOODS_NOMENCLATURE_DESCRIPTION_PERIOD_SID	= ws.cell(row = i, column = 5).value
	DATE				        				= ws.cell(row = i, column = 6).value

	if DATE == "":
		DATE = "30/03/2019"

	#print ("Here")
	obj = goods_nomenclature(GOODS_NOMENCLATURE_ITEM_ID, PRODUCTLINE_SUFFIX, DESCRIPTION, GOODS_NOMENCLATURE_SID, DATE, GOODS_NOMENCLATURE_DESCRIPTION_PERIOD_SID, "update", app)
	app.goods_nomenclature_list.append(obj)

"""
ws = wb['New']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	ADDITIONAL_CODE_TYPE_ID = ws.cell(row = i, column = 1).value
	DESCRIPTION             = ws.cell(row = i, column = 2).value
	APPLICATION_CODE        = ws.cell(row = i, column = 3).value

	f = additional_code_type(ADDITIONAL_CODE_TYPE_ID, DESCRIPTION, APPLICATION_CODE, "insert")
	app.additional_code_type_list.append(f)
"""
env = app.envelope_XML
env = env.replace("{ENVELOPE_ID}", str(app.base_envelope_id))
out = ""
for obj in app.goods_nomenclature_list:
	if 2 > 1:
	#if obj.still_exists:
		obj.writeXML(app)
		out += obj.xml

out = env.replace("{BODY}", out)
filename = os.path.join(app.XML_DIR, "goods_nomenclature.xml")
f = open(filename, "w", encoding="utf-8") 
f.write(out)
f.close()

schema_path = os.path.join(app.SCHEMA_DIR, "envelope.xsd")
my_schema = xmlschema.XMLSchema(schema_path)
try:
	if my_schema.is_valid(filename):
		print ("The file validated successfully.")
	else:
		print ("The file did not validate.")
except:
	print ("The file did not validate and crashed the validator.")

app.set_config()
app.set_config2("goods.nomenclature.description.periods", app.last_goods_nomenclature_description_period_sid)

