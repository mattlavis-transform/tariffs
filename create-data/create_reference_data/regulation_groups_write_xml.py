import psycopg2
import sys
import os
import xmlschema
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle

import common.objects as o
from common.regulation_group import regulation_group
from common.application import application

app = o.app
app.getTemplates()

try:
	arg = sys.argv[1]
except:
	arg = "regulation_groups"

fname = arg + ".xlsx"
fname2 = os.path.join(app.SOURCE_DIR, fname)
wb = load_workbook(filename = fname2, read_only = True)
ws = wb['Updated']
#print (fname2)
#sys.exit()

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	REGULATION_GROUP_ID = ws.cell(row = i, column = 1).value
	DESCRIPTION         = ws.cell(row = i, column = 2).value
	VALIDITY_START_DATE	= ws.cell(row = i, column = 3).value

	obj = regulation_group(REGULATION_GROUP_ID, DESCRIPTION, VALIDITY_START_DATE, "update")
	app.regulation_groups_list.append(obj)    

ws = wb['New']

row_count = ws.max_row
col_count = ws.max_column

for i in range(2, row_count + 1):
	REGULATION_GROUP_ID = ws.cell(row = i, column = 1).value
	DESCRIPTION         = ws.cell(row = i, column = 2).value
	VALIDITY_START_DATE	= ws.cell(row = i, column = 3).value

	obj = regulation_group(REGULATION_GROUP_ID, DESCRIPTION, VALIDITY_START_DATE, "insert")
	app.regulation_groups_list.append(obj)

env = app.envelope_XML
env = env.replace("{ENVELOPE_ID}", str(app.base_envelope_id))
out = ""
for obj in app.regulation_groups_list:
	obj.writeXML()
	out += obj.xml

out = env.replace("{BODY}", out)
fname3 = fname.replace("xlsx", "xml")
filename = os.path.join(app.XML_DIR, fname3)
print (filename)
f = open(filename, "w", encoding="utf-8") 
f.write(out)
f.close()
#sys.exit()
schema_path = os.path.join(app.SCHEMA_DIR, "envelope.xsd")
my_schema = xmlschema.XMLSchema(schema_path)
try:
	if my_schema.is_valid(filename):
		print ("The file validated successfully.")
		success = True
	else:
		print ("The file did not validate.")
		success = False
except:
	print ("The file did not validate and crashed the validator.")
	success = False

if not(success):
	my_schema.validate(filename)


app.set_config()
