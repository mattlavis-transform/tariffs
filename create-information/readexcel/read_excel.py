import psycopg2
import sys

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from datetime import datetime

def left_part(s, p):
	a = s.split(p)
	return a[0]

def muc_lookup(s):
	if s == "DTN":
		return ("/ 100 kg")
	elif s == "TNE":
		return ("/ 1000 kg")
	elif s == "kg E":
		return ("/ 100 kg/net eda")
	else:
		return (s)

def trap(x):
	try:
		if x is None:
			return("")
		else:
			return (x)
	except:
		return ("")

# Read in the measure XML template
f = open("C:\\projects\\readexcel\\templates\\measure.txt","r")
sTemplateMeasure = f.read()
# print (string)

		
wb = load_workbook(filename='C:\\projects\\writeexcel\\output\\R120978 GSP Generalised System of Preferences.xlsm', read_only=True)
ws = wb['working sheet']

row_count = ws.max_row
col_count = ws.max_column

"""
print(row_count)
print(col_count)
"""

sOut = ""

for i in range(3, row_count + 1):
	commodity_code     = ws.cell(row = i, column = 1).value.rjust(10, "0")
	sensitive_line     = ws.cell(row = i, column = 2).value
	reduction          = ws.cell(row = i, column = 3).value
	measure_type       = colon_split(ws.cell(row = i, column = 4).value, " : ")
	regulation         = ws.cell(row = i, column = 5).value
	start_date         = left_part(str(trap(ws.cell(row = i, column = 6).value)), *" ")
	end_date           = str(trap(ws.cell(row = i, column = 7).value)).
	geographical_area  = left_part(ws.cell(row = i, column = 8).value, " : ")
	ad_valorem         = ws.cell(row = i, column = 9).value
	specific           = ws.cell(row = i, column = 10).value
	specific_mu        = ws.cell(row = i, column = 11).value
	specific_muq       = ws.cell(row = i, column = 12).value
	ceiling            = ws.cell(row = i, column = 13).value
	ceiling_mu         = ws.cell(row = i, column = 14).value
	ceiling_muq        = ws.cell(row = i, column = 15).value
	meursing_agri      = ws.cell(row = i, column = 16).value
	meursing_sugar     = ws.cell(row = i, column = 17).value
	meursing_flour     = ws.cell(row = i, column = 18).value
	
	print(commodity_code)
	print(sensitive_line)
	print(reduction)
	print(measure_type)
	print(regulation)
	
	sBase = sTemplateMeasure
	sBase = sBase.replace("MEASURE_TYPE_ID",                  measure_type)
	sBase = sBase.replace("GEOGRAPHICAL_AREA_ID",             geographical_area)
	sBase = sBase.replace("GOODS_NOMENCLATURE_ITEM_ID",       commodity_code)
	sBase = sBase.replace("VALIDITY_START_DATE",              start_date)
	sBase = sBase.replace("VALIDITY_END_DATE",                end_date)
	sBase = sBase.replace("MEASURE_GENERATING_REGULATION_ID", regulation)
	sBase = sBase.replace("JUSTIFICATION_REGULATION_ID",      regulation)
	
	sOut += sBase

file = open("testfile.txt", "w")
file.write(sOut) 
file.close() 
		
def create_excel(regulation_id):
	# Connect to database
	conn = psycopg2.connect("dbname=trade_tariff_June18 user=postgres password=zanzibar")
	#regulation_id = "R120978"
	path = "C:\\projects\\writeexcel\\output\\"
	output_path = path + regulation_id + ".xlsx"

	# Visual parameters
	fill = PatternFill(fill_type = 'solid', start_color='000000') # , end_color='000000')
	fontMain   = Font(name='Calibri', size=9, bold=False, italic=False, color='000000')
	fontHeader = Font(name='Calibri', size=9, bold=True,  italic=False, color='ffffff')
	# Create a new workbook
	wb = Workbook()
	sheet=wb.active

	# Set up the table headers
	sheet['A1'] = "Measure SID"
	sheet['B1'] = "Measure type"
	sheet['C1'] = "Commodity"
	sheet['D1'] = "Additional code"
	sheet['E1'] = "Geography"
	sheet['F1'] = "Valid dates"
	sheet['G1'] = "Order number"
	sheet['H1'] = "Measure components"
	sheet['I1'] = "Conditions"
	sheet['J1'] = "Footnotes"

	# Get all the measures in the regulation
	sSQL = """SELECT v5.measure_sid, v5.regulation_id, v5.goods_nomenclature_item_id, v5.additional_code_type_id, v5.additional_code_id, v5.measure_type_id, v5.geographical_area_id, v5.validity_start_date, v5.validity_end_date, v5.ordernumber, mtd.description as measure_type_description, ga.description as geographical_area_description
	FROM v5, measure_type_descriptions mtd, ml_geographical_areas ga
	WHERE regulation_id = '""" + regulation_id + """'
	AND v5.measure_type_id = mtd.measure_type_id
	AND v5.geographical_area_id = ga.geographical_area_id ORDER BY measure_sid
	"""
	cur = conn.cursor()
	cur.execute(sSQL)
	rows_measures = cur.fetchall()
	measure_count = len(rows_measures)

	# Get its measure components
	sSQL = """SELECT mc.*, de.description as duty_expression_description, mucd.description as mucd_description
	FROM duty_expression_descriptions de, measure_components mc LEFT OUTER JOIN measurement_unit_qualifier_descriptions mucd ON mc.measurement_unit_qualifier_code = mucd.measurement_unit_qualifier_code
	WHERE measure_sid IN (SELECT measure_sid FROM v5 WHERE regulation_id = '""" + regulation_id + """')
	AND mc.duty_expression_id = de.duty_expression_id
	ORDER BY measure_sid, duty_expression_id
	"""

	cur = conn.cursor()
	cur.execute(sSQL)
	rows_components = cur.fetchall()
	component_count = len(rows_components)

	# Get footnotes
	sSQL = """SELECT fam.measure_sid, fam.footnote_type_id, fam.footnote_id, f.description FROM footnote_association_measures fam, ml_footnotes f
	WHERE f.footnote_id = fam.footnote_id
	AND f.footnote_type_id = fam.footnote_type_id
	AND fam.measure_sid IN (SELECT measure_sid FROM v5 WHERE regulation_id = '""" + regulation_id + """')
	ORDER BY fam.measure_sid, fam.footnote_type_id, fam.footnote_id
	"""
	cur = conn.cursor()
	cur.execute(sSQL)
	rows_footnotes = cur.fetchall()
	footnote_count = len(rows_footnotes)

	# Get conditions

	sSQL = """SELECT mc.measure_sid, mc.measure_condition_sid, mc.condition_code, mc.condition_duty_amount, 
	mc.condition_measurement_unit_code, mc.condition_measurement_unit_qualifier_code, mc.action_code, mc.certificate_type_code, mc.certificate_code,
	mccd.description as condition_code_description, mad.description as action_code_description, cc.description as certificate_description
	FROM 
	measure_action_descriptions mad, measure_condition_code_descriptions mccd, ml_certificate_codes cc
	RIGHT OUTER JOIN measure_conditions mc ON cc.certificate_code = mc.certificate_code AND cc.certificate_type_code = mc.certificate_type_code
	WHERE mccd.condition_code = mc.condition_code
	AND mad.action_code = mc.action_code
	AND mc.measure_sid IN (SELECT measure_sid FROM v5 WHERE regulation_id = '""" + regulation_id + """')
	ORDER BY mc.measure_sid, mc.component_sequence_number
	"""

	cur = conn.cursor()
	cur.execute(sSQL)
	rows_conditions = cur.fetchall()
	condition_count = len(rows_conditions)

	# Get condition components
	sSQL = """SELECT mc.measure_sid, mcc.measure_condition_sid, mcc.duty_expression_id, mcc.duty_amount, mcc.monetary_unit_code,
	mcc.measurement_unit_code, mcc.measurement_unit_qualifier_code, ded.description as duty_expression_description
	FROM measure_conditions mc, measure_condition_components mcc, duty_expression_descriptions ded
	WHERE mc.measure_condition_sid = mcc.measure_condition_sid
	AND mcc.duty_expression_id = ded.duty_expression_id
	AND mc.measure_sid IN (SELECT measure_sid FROM v5 WHERE regulation_id = '""" + regulation_id + """')
	ORDER BY 1, 2, 3"""

	cur = conn.cursor()
	cur.execute(sSQL)
	rows_condition_components = cur.fetchall()
	condition_component_count = len(rows_condition_components)

	lRow   = 2
	lStart = 0
	lStartF = 0
	lStartC = 0
	lStartCC = 0

	for rd in rows_measures:
		measure_sid = rd[0]
		additional_code_type_id = trap(rd[3])
		additional_code_id = trap(rd[4])
		validity_start_date = str(trap(rd[7]))
		validity_end_date = str(trap(rd[8]))

		pos = validity_start_date.find(" ")
		validity_start_date2 = validity_start_date[0:pos]
		
		if validity_end_date == "":
			validity_end_date2 = "[open-ended]"
		else:
			pos = validity_end_date.find(" ")
			validity_end_date2 = validity_end_date[0:pos]
		
		sheet.cell(row = lRow, column = 1).value = measure_sid
		sheet.cell(row = lRow, column = 2).value = "[" + rd[5] + "] " + rd[10]
		sheet.cell(row = lRow, column = 3).value = rd[2]
		sheet.cell(row = lRow, column = 4).value = additional_code_type_id = additional_code_id
		sheet.cell(row = lRow, column = 5).value = "[" + rd[6] + "] " + rd[11]
		sheet.cell(row = lRow, column = 6).value = validity_start_date2 + ' - ' + validity_end_date2
		sheet.cell(row = lRow, column = 7).value = trap(rd[9])

		# Get duties
		sDuty = ""
		sDutyExpressionList = ""
		for x in range(lStart, component_count):
			rd2 = rows_components[x]
			measure_sid2                    = rd2[0]
			duty_expression_id              = rd2[1]
			#duty_amount           = "{:12.2f}".format(rd2[2])
			duty_amount                     = str(rd2[2])
			monetary_unit_code              = str(rd2[3])
			measurement_unit_code           = trap(rd2[4])
			measurement_unit_qualifier_code = trap(rd2[5])
			description                     = rd2[9]
			mucd_description                = rd2[10]
			if duty_amount == "None":
				duty_amount = ""
			if measure_sid == measure_sid2:
				sDutyExpressionList += "[" + str(duty_expression_id) + "]"
				
				if description == "Maximum":
					description = " up to a maximum of "
				elif description == "Minimum":
					description = " to a minimum of "
				if description.find("% or amount") > -1:
					description = description.replace("% or amount", "")
					if monetary_unit_code != "None":
						sDuty += description + " " + duty_amount + " " + monetary_unit_code + " "
					else:
						sDuty += description + " " + duty_amount + "% "
				else:
					if duty_expression_id in ['15', '17', '35']:
						sDuty += description + str(duty_amount) + " "
					else:
						sDuty += str(duty_amount) + " " + description + " "
					if monetary_unit_code != "None":
						sDuty += monetary_unit_code + " "

				if measurement_unit_code != "":
					sDuty += " " + muc_lookup(measurement_unit_code) + " "
				if measurement_unit_qualifier_code != "":
					sDuty += " " + mucd_description + " "
			else:
				lStart = x
				break
		sDuty = sDuty.replace("  ", " ")
		sDuty.rstrip()
		sheet.cell(row = lRow, column = 8).value = sDuty + " " + sDutyExpressionList
		# End get duty
		
		# Start get conditions
		sConditionString = ""
		my_condition_count = 0
		measure_condition_sid_last = -1
		for x in range(lStartC, condition_count):
			rd3 = rows_conditions[x]
			measure_sid2                              = trap(rd3[0])
			measure_condition_sid                     = trap(rd3[1])
			condition_code                            = trap(rd3[2])
			condition_duty_amount                     = trap(rd3[3])
			condition_measurement_unit_code           = trap(rd3[4])
			condition_measurement_unit_qualifier_code = trap(rd3[5])
			action_code                               = trap(rd3[6])
			certificate_type_code                     = trap(rd3[7])
			certificate_code                          = trap(rd3[8])
			condition_code_description                = trap(rd3[9])
			action_code_description                   = trap(rd3[10])
			certificate_description                   = trap(rd3[11])
			if measure_sid == measure_sid2:
				my_condition_count += 1
				if my_condition_count != 1:
					sConditionString += "\n"
				sConditionString += "Condition [" + condition_code + str(my_condition_count) + "] " + condition_code_description + "\n"
				if certificate_type_code != "":
					sConditionString += "Document [" + certificate_type_code + certificate_code + "] " + certificate_description + "\n"
				sConditionString += "Action [" + action_code + "] " + action_code_description + "\n"
				if condition_duty_amount != "":
					sConditionString += "Requirement - " + str(condition_duty_amount) + " " + condition_measurement_unit_code + " " + condition_measurement_unit_qualifier_code + "\n"
				# Need to loop around on the same measure_condition_sid

				my_component_count = 0
				duty_added = False
				for x2 in range(lStartCC, condition_component_count):
					rd4 = rows_condition_components[x2]
					measure_sid3                    = trap(rd4[0])
					measure_condition_sid2          = trap(rd4[1])
					duty_expression_id              = trap(rd4[2])
					duty_amount                     = trap(rd4[3])
					monetary_unit_code              = trap(rd4[4])
					measurement_unit_code           = trap(rd4[5])
					measurement_unit_qualifier_code = trap(rd4[6])
					duty_expression_description     = trap(rd4[7])
					if (measure_sid3 == measure_sid) and (measure_condition_sid2 == measure_condition_sid):
						duty_added = True
						if my_component_count == 0:
							sConditionString += "Applicable duty: "
						my_component_count += 1
						sConditionString += "[" + duty_expression_id + "] " + str(duty_amount) + duty_expression_description + " " + monetary_unit_code + " " + measurement_unit_code + " " + measurement_unit_qualifier_code + " "
					else:
						if duty_added == True:
							sConditionString += "\n"
						lStartCC = x2
						break
					
			else:
				lStartC = x
				break
			measure_condition_sid_last = measure_condition_sid
		sConditionString = sConditionString.replace("  ", " ")
		sConditionString.rstrip()
		sheet.cell(row = lRow, column = 9).value = sConditionString
		# End get conditions
		
		# Start get footnotes
		sFootnoteString = ""
		for x in range(lStartF, footnote_count):
			rd2 = rows_footnotes[x]
			measure_sid2     = rd2[0]
			footnote_type_id = rd2[1]
			footnote_id      = rd2[2]
			description      = rd2[3]
			if measure_sid == measure_sid2:
				sFootnoteString += footnote_type_id + footnote_id + " " + description + "\n"
			else:
				lStartF = x
				break
		sFootnoteString = sFootnoteString.replace("  ", " ")
		sFootnoteString.rstrip()
		sheet.cell(row = lRow, column = 10).value = sFootnoteString
		# End get footnotes

		lRow += 1

	# freeze panes
	c = sheet['A2']
	sheet.freeze_panes = c
		
	# Set fixed widths on columns
	sheet.column_dimensions["A"].width = 15
	sheet.column_dimensions["B"].width = 30
	sheet.column_dimensions["C"].width = 20
	sheet.column_dimensions["D"].width = 20
	sheet.column_dimensions["E"].width = 50
	sheet.column_dimensions["F"].width = 30
	sheet.column_dimensions["G"].width = 15
	sheet.column_dimensions["H"].width = 60
	sheet.column_dimensions["I"].width = 75
	sheet.column_dimensions["J"].width = 75

	# cell wrap
	for i, rowOfCellObjects in enumerate(sheet['A1':'J' + str(measure_count + 1)]):
		for n, cellObj in enumerate(rowOfCellObjects):
			cellObj.alignment = Alignment(wrapText=True, horizontal='left', vertical='top')
			cellObj.font = fontMain
			
	# Apply styling to the table headers
	for i, rowOfCellObjects in enumerate(sheet['A1':'J1']):
		for n, cellObj in enumerate(rowOfCellObjects):
			cellObj.fill = fill
			cellObj.font = fontHeader

	# Add a filter
	sheet.auto_filter.ref = "A:J"

	wb.save(output_path)