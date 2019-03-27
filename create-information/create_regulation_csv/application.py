import psycopg2
import os
import csv
import sys
import codecs
import re
import functions
import xlsxwriter

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from shutil import copyfile
from collections import Counter

from measure	import measure
from component	import component
from footnote	import footnote
from condition	import condition
from quota		import quota
from commodity	import commodity

class application(object):
	def __init__(self):
		self.conn = psycopg2.connect("dbname=trade_tariff_181212b user=postgres password" + self.p)
		#self.conn = psycopg2.connect("dbname=trade_tariff_181119 user=postgres password" + self.p)
		self.BASE_DIR				= os.path.dirname(os.path.abspath(__file__))
		self.EXCEL_DIR				= os.path.join(self.BASE_DIR,		"excel")
		self.COMMODITY_DIR			= os.path.join(self.BASE_DIR,		"commodities")
		self.AGREEMENT_DIR			= os.path.join(self.BASE_DIR,		"agreements")
		self.QUOTA_WORD_DIR			= os.path.join(self.BASE_DIR,		"..")
		self.QUOTA_WORD_DIR			= os.path.join(self.QUOTA_WORD_DIR,	"create_quota_sheet")
		self.QUOTA_WORD_DIR			= os.path.join(self.QUOTA_WORD_DIR,	"word")

		self.lst_preferential_measures = ['142', '143', '145', '146']
		
		self.write_csv				= False
		self.reinitialise()
		self.readCommodities()
		self.readMeasureTypeDescriptions()
		self.readRegulationGroupDescriptions()

	def reinitialise(self):
		self.meursing_count			= 0
		self.siv_count				= 0
		self.footnote_count			= 0
		self.condition_count		= 0
		self.siv_measure_sid_list	= set()
		self.footnote_list			= set()
		self.quota_set				= set()

	def createRegulationCSV(self, sRegulationID):
		###############################################################
		# Get the measures
		sSQL = """SELECT mtd.description AS mt, m.measure_type_id as mtid, m.measure_sid, m.geographical_area_id AS geo,
		m.goods_nomenclature_item_id, m.additional_code_type_id, m.additional_code_id, m.validity_start_date,
		m.validity_end_date, m.ordernumber, mlac.description AS adcd, mlga.description as gd, m.regulation_id_full, m.regulation_group_id
		FROM measure_type_descriptions mtd, ml.v5_2019 m LEFT OUTER JOIN ml.ml_additional_codes mlac
		ON (m.additional_code_type_id = mlac.additional_code_type_id AND m.additional_code_id = mlac.additional_code)
		LEFT OUTER JOIN ml.ml_geographical_areas mlga ON (m.geographical_area_id = mlga.geographical_area_id)
		WHERE mtd.measure_type_id = m.measure_type_id AND m.regulation_id = '""" + sRegulationID + """'
		ORDER BY m.goods_nomenclature_item_id, m.validity_start_date;
		"""
		#print (sSQL)
		#sys.exit()
		cur = self.conn.cursor()
		cur.execute(sSQL)
		rows_measures = cur.fetchall()
		lstMeasures		= []
		lstMeasureTypes	= []
		self.lstRegulations	= []
		sInClause = ""
		
		# Get the measures
		for m in rows_measures:
			measure_type_description		= m[0]
			measure_type_id					= m[1]
			measure_sid						= m[2]
			geographical_area_id			= m[3]
			goods_nomenclature_item_id		= m[4]
			sAdditionalCodeTypeID			= m[5]
			sAdditionalCodeID				= m[6]
			sValidityStartDate				= m[7]
			sValidityEndDate				= m[8]
			order_number					= functions.mstr(m[9])
			additional_code_description		= m[10]
			geographical_area_description	= m[11]
			regulation_id_full				= m[12]
			regulation_group_id				= m[13]
			regulation_group_descriptor		= regulation_id_full + " - " + regulation_group_id
			self.lstRegulations.append(regulation_group_descriptor)

			oMeasure = measure(measure_type_description, measure_type_id, measure_sid, geographical_area_id, goods_nomenclature_item_id, sAdditionalCodeTypeID, sAdditionalCodeID, sValidityStartDate, sValidityEndDate, order_number, additional_code_description, geographical_area_description, regulation_id_full)
			lstMeasures.append(oMeasure)
			sInClause += str(measure_sid) + ", "
			lstMeasureTypes.append(measure_type_id)
			if order_number != "":
				self.quota_set.add (order_number)
			
		self.lstMeasureTypesUnique = Counter(lstMeasureTypes)
		self.lstRegulationsUnique = Counter(self.lstRegulations)
		self.dictMeasureTypesUnique = dict(self.lstMeasureTypesUnique)
		#print (self.lstRegulationsUnique)
		#sys.exit()
		self.nonpreferential_count = 0
		for d in self.lstMeasureTypesUnique:
			if d not in self.lst_preferential_measures:
				self.nonpreferential_count += self.dictMeasureTypesUnique[d]


		sInClause = sInClause.strip(", ")

		# Get the measure components (duties)
		sSQL = """SELECT mc.measure_sid, mc.duty_amount, mc.monetary_unit_code as muc, mc.duty_expression_id, mud.measurement_unit_code,
		muqd.measurement_unit_qualifier_code, ded.description AS ded FROM measure_components mc LEFT OUTER JOIN duty_expression_descriptions ded
		on (mc.duty_expression_id = ded.duty_expression_id) LEFT OUTER JOIN measurement_unit_descriptions mud on
		(mc.measurement_unit_code = mud.measurement_unit_code) LEFT OUTER JOIN measurement_unit_qualifier_descriptions muqd
		on (mc.measurement_unit_qualifier_code = muqd.measurement_unit_qualifier_code)
		WHERE mc.measure_sid IN ( """ + sInClause + """) ORDER BY 1, mc.duty_expression_id;
		"""
		cur = self.conn.cursor()
		cur.execute(sSQL)
		rows_measure_components = cur.fetchall()
		lstMeasureComponents = []

		for mc in rows_measure_components:
			measure_sid								= mc[0]
			duty_amount								= mc[1]
			monetary_unit_code						= mc[2]
			duty_expression_id						= mc[3]
			measurement_unit_code					= mc[4]
			measurement_unit_qualifier_code			= mc[5]
			oMeasureComponent = component(measure_sid, duty_amount, monetary_unit_code, duty_expression_id, measurement_unit_code, measurement_unit_qualifier_code, "Active")
			lstMeasureComponents.append(oMeasureComponent)

		# Get the footnotes
		sSQL = """SELECT DISTINCT fam.measure_sid, fam.footnote_type_id, fam.footnote_id, fd.description as fd, ftd.description AS ftd
		FROM footnote_association_measures fam, ml.ml_footnotes fd, footnote_type_descriptions ftd
		WHERE fd.footnote_type_id = fam.footnote_type_id AND fd.footnote_id = fam.footnote_id AND ftd.footnote_type_id = fam.footnote_type_id
		AND fam.measure_sid IN (""" + sInClause + """) ORDER BY 1, 2, 3;"""
		cur = self.conn.cursor()
		cur.execute(sSQL)
		rows_footnotes = cur.fetchall()
		lstFootnotes = []
		for r in rows_footnotes:
			measure_sid              = r[0]
			sFootnoteTypeID          = r[1]
			sFootnoteID              = r[2]
			sFootnoteDescription     = r[3]
			sFootnoteTypeDescription = r[4]
			
			oFootnote = footnote(measure_sid, sFootnoteTypeID, sFootnoteID, sFootnoteDescription, sFootnoteTypeDescription)
			lstFootnotes.append(oFootnote)


		# Get the conditions
		sSQL = """SELECT mc.measure_sid, mc.measure_condition_sid, mc.condition_code, mccd.description as mccd, mc.action_code, mad.description as mad, mc.condition_duty_amount,
		mc.certificate_type_code, ctd.description as ctd, mc.certificate_code, cd.description as cd, mc.condition_monetary_unit_code, mc.condition_measurement_unit_code,
		mc.condition_measurement_unit_qualifier_code, mud.description as mud, muqd.description muqd
		FROM measure_conditions mc LEFT OUTER JOIN measure_action_descriptions mad
		ON mc.action_code = mad.action_code LEFT OUTER JOIN measure_condition_code_descriptions mccd ON mc.condition_code = mccd.condition_code
		LEFT OUTER JOIN certificate_type_descriptions ctd ON mc.certificate_type_code = ctd.certificate_type_code LEFT OUTER JOIN ml.ml_certificate_codes cd
		ON (mc.certificate_code = cd.certificate_code AND mc.certificate_type_code = cd.certificate_type_code) LEFT OUTER JOIN measurement_unit_descriptions mud
		ON mc.condition_measurement_unit_code = mud.measurement_unit_code LEFT OUTER JOIN measurement_unit_qualifier_descriptions muqd
		ON mc.condition_measurement_unit_qualifier_code = muqd.measurement_unit_qualifier_code WHERE measure_sid IN (""" + sInClause + """)
		ORDER BY mc.measure_sid, mc.component_sequence_number;"""
		cur = self.conn.cursor()
		cur.execute(sSQL)
		rows_conditions = cur.fetchall()
		lstConditions = []
		for r in rows_conditions:
			measure_sid                              = r[0]
			sMeasureConditionSID                     = r[1]
			sConditionCode                           = r[2]
			sConditionCodeDescription                = r[3]
			sActionCode                              = r[4]
			sMeasureActionDescription                = r[5]
			sConditionDutyAmount                     = r[6]
			sCertificateTypeCode                     = r[7]
			sCertificateTypeCodeDescription          = r[8]
			sCertificateCode                         = r[9]
			sCertificateDescription                  = r[10]
			monetary_unit_code                       = r[11]
			sMeasurementUnitCode                     = r[12]
			sMeasurementUnitQualifierCode            = r[13]
			sMeasurementUnitCodeDescription          = r[14]
			sMeasurementUnitQualifierCodeDescription = r[15]
			
			oCondition = condition(measure_sid, sMeasureConditionSID, sConditionCode, sConditionCodeDescription, sActionCode, sMeasureActionDescription, sConditionDutyAmount, sCertificateTypeCode, sCertificateTypeCodeDescription, sCertificateCode, sCertificateDescription, monetary_unit_code, sMeasurementUnitCode, sMeasurementUnitQualifierCode, sMeasurementUnitCodeDescription, sMeasurementUnitQualifierCodeDescription)
			
			lstConditions.append(oCondition)


		for m in lstMeasures:
			m.addComponents(lstMeasureComponents)
			m.addFootnotes(lstFootnotes)
			m.addConditions(lstConditions)

		# Transfer quota set into quota list
		self.quota_list = []
		self.licensed_quota_count = 0
		for q in self.quota_set:
			q2 = quota(q)
			self.quota_list.append(q2)
			if q[:3] == "094":
				self.licensed_quota_count += 1

		
		for m in lstMeasures:
			if m.order_number != "":
				for q in self.quota_list:
					if m.order_number == q.order_number:
						q.measure_count += 1
						q.measure_type = m.measure_type_description
						c = commodity(m.goods_nomenclature_item_id, m.goods_description)
						q.commodity_list.append (c)
						q.date_set.add (m.validity_dates)

		if len(self.quota_list) > 0:
			for q in self.quota_list:
				q.get_specials()
			
		# Write the CSV file
		if self.write_csv:
			sFilename = "csv\\" + sRegulationID + ".csv"
			with open(sFilename, 'w', newline='', encoding='utf-8') as csvfile:
				filewriter = csv.writer(csvfile, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)
				filewriter.writerow(["SID", "Measure type", "Commodity", "Additional code", "Geography", "Valid dates", "Order number", "Measure components", "Conditions", "Footnotes"])

				for m in lstMeasures:
					filewriter.writerow([m.measure_sid, m.measure_type_full, m.goods_nomenclature_item_id, m.additional_code_full, m.validity_dates, m.order_number, m.components, m.conditions, m.footnotes])
	
		# Write the Excel file
		self.excel_file = os.path.join(self.EXCEL_DIR, sRegulationID + ".xlsx")
		
		workbook = xlsxwriter.Workbook(self.excel_file)
		worksheet = workbook.add_worksheet()
		worksheet.name = "Raw"
		bold = workbook.add_format({'bold': True, 'font_color': 'white', 'bg_color': 'black'})
		bold.set_font_size(9)
		
		hd = workbook.add_format({'bold': True, 'font_color': 'black'})
		hd.set_font_size(9)
		
		cfWrap = workbook.add_format()
		cfWrap.set_text_wrap()
		cfWrap.set_align('left')
		cfWrap.set_align('top')
		cfWrap.set_font_size(9)
		cfWrap.set_font_name('Calibri')
		
		worksheet.write('A1', 'Measure SID', bold)
		worksheet.write('B1', 'Measure type', bold)
		worksheet.write('C1', 'Commodity', bold)
		worksheet.write('D1', 'Additional code', bold)
		worksheet.write('E1', 'Geography', bold)
		worksheet.write('F1', 'Valid dates', bold)
		worksheet.write('G1', 'Order number', bold)
		worksheet.write('H1', 'Measure components', bold)
		worksheet.write('I1', 'Conditions', bold)
		worksheet.write('J1', 'Footnotes', bold)

		worksheet.set_column("A:A", 12)
		worksheet.set_column("B:B", 25)
		worksheet.set_column("C:D", 15)
		worksheet.set_column("E:E", 20)
		worksheet.set_column("F:F", 20)
		worksheet.set_column("G:G", 12)
		worksheet.set_column("H:H", 32)
		worksheet.set_column("I:I", 80)
		worksheet.set_column("J:J", 60)

		i = 1
		my_dates = set()
		for m in lstMeasures:
			worksheet.write(i, 0, m.measure_sid, cfWrap)
			worksheet.write(i, 1, m.measure_type_full, cfWrap)
			worksheet.write(i, 2, m.goods_nomenclature_item_id, cfWrap)
			worksheet.write(i, 3, m.additional_code_full, cfWrap)
			worksheet.write(i, 4, m.geographical_area_full, cfWrap)
			worksheet.write(i, 5, m.validity_dates, cfWrap)
			worksheet.write(i, 6, m.order_number, cfWrap)
			worksheet.write(i, 7, m.components_excel, cfWrap)
			worksheet.write(i, 8, m.conditions_excel, cfWrap)
			worksheet.write(i, 9, m.footnotes, cfWrap)
			i = i + 1



		worksheet.freeze_panes(1, 0)
		worksheet.autofilter("A1:J1")

		# Write worksheet 2 with the summaries
		worksheet2 = workbook.add_worksheet()
		worksheet2.name = "Stats"

		worksheet2.write('A1', 'Item',  bold)
		worksheet2.write('B1', 'Detail', bold)
		worksheet2.set_column("A:A",    40)
		worksheet2.set_column("B:B",    30)
		
		line = 1
		# Total measure count
		worksheet2.write(line, 0, "Total measure count",  cfWrap)
		worksheet2.write(line, 1, str(len(lstMeasures)), cfWrap)
		line += 1
		# Write quota count
		worksheet2.write(line, 0, "Total quota count",  cfWrap)
		worksheet2.write(line, 1, str(len(self.quota_list)), cfWrap)
		line += 1
		# Write quota count
		worksheet2.write(line, 0, "Licensed quota count",  cfWrap)
		worksheet2.write(line, 1, str(self.licensed_quota_count), cfWrap)
		line += 1
		# Write SIVs
		worksheet2.write(line, 0, "SIV count",  cfWrap)
		worksheet2.write(line, 1, str(self.siv_count), cfWrap)
		line += 1
		# Write Meursing
		worksheet2.write(line, 0, "Meursing reference count",  cfWrap)
		worksheet2.write(line, 1, str(self.meursing_count), cfWrap)
		line += 1
		# Write footnotes
		worksheet2.write(line, 0, "Footnote count",  cfWrap)
		worksheet2.write(line, 1, str(self.footnote_count), cfWrap)
		line += 1
		# Write conditions
		worksheet2.write(line, 0, "Condition count",  cfWrap)
		worksheet2.write(line, 1, str(self.condition_count), cfWrap)
		line += 2
		
		# Write regulations heading
		worksheet2.write(line, 0, "Regulations",  hd)
		line += 1
		# Write regulation count
		worksheet2.write(line, 0, "Regulation parts",  cfWrap)
		self.regulation_count = 0

		for reg in self.lstRegulationsUnique:
			s = reg.split("-")
			grp = s[1].strip(" ")
			worksheet2.write(line, 0, reg + " (" + self.dict_regulation_group_descriptions.get(grp) + ")",  cfWrap)
			worksheet2.write(line, 1, self.lstRegulationsUnique[reg], cfWrap)
			line += 1

		line += 1

		# Write measures heading
		worksheet2.write(line, 0, "Measure types",  hd)
		line += 1

		# Write non-preferential measure type
		worksheet2.write(line, 0, "Non-preferential measure count",  cfWrap)
		worksheet2.write(line, 1, str(self.nonpreferential_count), cfWrap)
		line += 1

		# Write list of all measure type
		for mt in self.lstMeasureTypesUnique:
			measure_type_description = self.dict_measure_type_descriptions[mt]
			worksheet2.write(line, 0, measure_type_description + " [" + mt + "]",  cfWrap)
			worksheet2.write(line, 1, self.dictMeasureTypesUnique[mt], cfWrap)
			line += 1

		# Write worksheet 3 with the quotas
		worksheet3 = workbook.add_worksheet()
		worksheet3.name = "Quotas"

		worksheet3.write('A1', 'Order number',  	bold)
		worksheet3.write('B1', 'Measure type',		bold)
		worksheet3.write('C1', 'Commodity codes (HS6)',	bold)
		worksheet3.write('D1', 'Periods captured in measures',	bold)
		worksheet3.write('E1', 'Periods - points of interest',	bold)
		
		worksheet3.set_column("A:A", 15)
		worksheet3.set_column("B:B", 30)
		worksheet3.set_column("C:C", 100)
		worksheet3.set_column("D:D", 30)
		worksheet3.set_column("E:E", 70)

		i = 1

		for q in self.quota_list:
			worksheet3.write(i, 0, q.order_number, cfWrap)
			worksheet3.write(i, 1, q.measure_type, cfWrap)
			if len(q.commodity_list) > 0:
				s = ""
				commodity_count = 0
				for c in q.commodity_list:
					s += c.commodity_code + " : " + c.description + "...\n"
					commodity_count += 1
				worksheet3.write(i, 2, s.strip("\n"), cfWrap)
			if len(q.date_set) > 0:
				s = ""
				for d in sorted(q.date_set):
					s += d + "\n"
				worksheet3.write(i, 3, s.strip("\n"), cfWrap)
			worksheet3.write(i, 4, q.special_text, cfWrap)
			i += 1

		worksheet3.freeze_panes(1, 0)
		worksheet3.autofilter("A1:E1")

		workbook.close()

		self.copy_documents(sRegulationID)
		
	def create_agreement(self, my_agreement):
		self.agreement_name	= my_agreement
		self.agreement_path	= os.path.join(self.AGREEMENT_DIR, my_agreement)
		self.quota_path		= os.path.join(self.agreement_path, "quotas")

		try:
			os.mkdir(self.AGREEMENT_DIR)
		except:
			pass
		try:
			os.mkdir(self.agreement_path)
		except:
			pass
		try:
			os.mkdir(self.quota_path)
		except:
			pass
	
	def copy_documents(self, sRegulationID):
		try:
			src		= os.path.join(self.EXCEL_DIR,		self.excel_file)
			dest	= os.path.join(self.agreement_path,	self.agreement_name + "_" + sRegulationID + ".xlsx")
		
			#print ("EF", self.excel_file)
			#print ("AP", self.agreement_path)
			#print ("SRC", src)
			#print ("DEST", dest)
			#sys.exit()
			copyfile(src, dest)
		except:
			pass
		
		for q in self.quota_list:
			filename = q.order_number + ".docx"
			src = os.path.join(self.QUOTA_WORD_DIR, filename)
			dest = os.path.join(self.quota_path, filename)
			try:
				copyfile(src, dest)
			except:
				pass


	def readCommodities(self):
		self.commodity_list = []
		filename = os.path.join(self.COMMODITY_DIR,	"hs6_nomenclature.csv")
		with open(filename) as csv_file:
			csv_reader = csv.reader(csv_file, delimiter=';')
			line_count = 0
			for row in csv_reader:
				if line_count == 0:
					pass
				else:
					commodity_code	= row[0]
					description		= row[1]
					c = commodity(commodity_code, description)
					self.commodity_list.append(c)
				line_count += 1

	def readMeasureTypeDescriptions(self):
		sql = """SELECT mt.measure_type_id, mtd.description FROM measure_types mt, measure_type_descriptions mtd
		WHERE mt.measure_type_id = mtd.measure_type_id
		AND validity_end_date IS NULL AND mt.measure_type_id < 'A'
		ORDER BY 1"""
		cur = self.conn.cursor()
		cur.execute(sql)
		rows = cur.fetchall()
		self.dict_measure_type_descriptions = dict(rows)


	def readRegulationGroupDescriptions(self):
		sql = """SELECT regulation_group_id, description FROM regulation_group_descriptions ORDER BY 1"""
		cur = self.conn.cursor()
		cur.execute(sql)
		rows = cur.fetchall()
		self.dict_regulation_group_descriptions = dict(rows)

